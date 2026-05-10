from pathlib import Path
from pypdf import PdfReader
from docx import Document


# Source-code and structured-text extensions. Treated as plain UTF-8 text;
# the LLM gets a dedicated 'code summarizer' persona for these.
CODE_EXT = frozenset(
    {
        # mainstream languages
        ".py", ".js", ".jsx", ".ts", ".tsx", ".mjs", ".cjs",
        ".java", ".kt", ".scala", ".groovy",
        ".c", ".h", ".cpp", ".cxx", ".cc", ".hpp", ".hxx",
        ".cs", ".m", ".mm",
        ".rs", ".go", ".rb", ".swift",
        ".php", ".pl", ".lua", ".r", ".sql",
        # shell + automation
        ".sh", ".bash", ".zsh", ".fish",
        # web / markup
        ".html", ".htm", ".css", ".scss", ".sass", ".less",
        # data / config (parsed as text — same prompt as code)
        ".json", ".yaml", ".yml", ".toml", ".ini", ".xml",
        ".csv", ".tsv", ".log",
    }
)


def _read_txt(p: Path) -> str:
    return p.read_text(encoding="utf-8", errors="ignore")


def _read_pdf(p: Path) -> str:
    reader = PdfReader(str(p))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def _read_docx(p: Path) -> str:
    doc = Document(str(p))
    return "\n".join(par.text for par in doc.paragraphs)


def _read_xlsx(p: Path) -> str:
    """Stringify every cell in every sheet, tab-separated. data_only=True
    so we get computed values instead of formula text."""
    from openpyxl import load_workbook  # lazy: openpyxl isn't tiny

    wb = load_workbook(filename=str(p), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in wb.worksheets:
        parts.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            # Skip fully-empty rows so the LLM doesn't waste tokens on whitespace.
            if any(c.strip() for c in cells):
                parts.append("\t".join(cells))
    return "\n".join(parts)


def is_code_path(path: Path) -> bool:
    return path.suffix.lower() in CODE_EXT


def parse_file(path: Path) -> str:
    ext = path.suffix.lower()
    # Document formats first
    if ext in {".txt", ".md", ".markdown", ".rtf"}:
        return _read_txt(path)
    if ext == ".pdf":
        return _read_pdf(path)
    if ext == ".docx":
        return _read_docx(path)
    if ext == ".xlsx":
        return _read_xlsx(path)
    # Source code + structured text — read as UTF-8 plain text
    if ext in CODE_EXT:
        return _read_txt(path)
    return ""
